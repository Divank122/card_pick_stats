import argparse
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_pick_rate_from_workbook(wb, card_id):
    for sheet_name in wb.sheetnames:
        if sheet_name == "卡组构成":
            continue
        ws = wb[sheet_name]
        for row_idx in range(3, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == card_id:
                total_rate = ws.cell(row=row_idx, column=14).value
                return total_rate if total_rate else 0
    return None

def compare_excels(file1, file2, output_file=None):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    
    if output_file is None:
        output_file = Path(file1).parent / f"comparison_{Path(file1).stem}_vs_{Path(file2).stem}.xlsx"
    else:
        output_file = Path(output_file)
    
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    
    for sheet_name in wb1.sheetnames:
        if sheet_name == "卡组构成":
            continue
        
        ws1 = wb1[sheet_name]
        
        ws_out = wb_out.create_sheet(title=sheet_name)
        
        header_row = ["卡牌ID", "卡牌", "抓取率", "差异"]
        ws_out.append(header_row)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 5):
            cell = ws_out.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx in range(3, ws1.max_row + 1):
            card_id = ws1.cell(row=row_idx, column=1).value
            if not card_id:
                continue
            
            card_name = ws1.cell(row=row_idx, column=2).value
            rate1 = ws1.cell(row=row_idx, column=14).value
            rate1 = rate1 if rate1 else 0
            
            rate2 = get_pick_rate_from_workbook(wb2, card_id)
            
            if rate2 is not None:
                diff = (rate1 - rate2) * 100
                if diff > 0:
                    diff_str = f"↑{diff:.1f}%"
                    diff_font = Font(color="008000", bold=(diff > 20))
                elif diff < 0:
                    diff_str = f"↓{abs(diff):.1f}%"
                    diff_font = Font(color="FF0000", bold=(abs(diff) > 20))
                else:
                    diff_str = "→0.0%"
                    diff_font = Font(color="000000", bold=False)
            else:
                diff_str = "N/A"
                diff_font = Font(color="808080")
            
            row_data = [card_id, card_name, rate1, diff_str]
            ws_out.append(row_data)
            
            new_row = ws_out.max_row
            ws_out.cell(row=new_row, column=3).number_format = '0.0%'
            ws_out.cell(row=new_row, column=4).font = diff_font
            ws_out.cell(row=new_row, column=4).alignment = Alignment(horizontal="center")
        
        ws_out.column_dimensions['A'].width = 25
        ws_out.column_dimensions['B'].width = 15
        ws_out.column_dimensions['C'].width = 10
        ws_out.column_dimensions['D'].width = 10
    
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_file)
    print(f"\n比较结果已保存到: {output_file}")
    return output_file

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="比较两个卡牌统计Excel文件的抓取率差异")
    parser.add_argument("file1", type=str, help="第一个Excel文件路径（作为基准）")
    parser.add_argument("file2", type=str, help="第二个Excel文件路径")
    parser.add_argument("--output", "-o", type=str, default=None,
                        help="输出文件路径（可选）")
    args = parser.parse_args()
    
    compare_excels(args.file1, args.file2, args.output)
