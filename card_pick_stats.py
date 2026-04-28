import json
import os
import argparse
from datetime import datetime
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

RARITY_NAMES = {
    "Basic": "基础",
    "Common": "普通",
    "Uncommon": "罕见",
    "Rare": "稀有",
    "Ancient": "远古",
    "Event": "事件",
    "Token": "代币",
    "Status": "状态",
    "Curse": "诅咒",
    "Quest": "任务",
    "Unknown": "未知"
}

RARITY_ORDER = ["Common", "Uncommon", "Rare", "Ancient", "Basic", "Event", "Token", "Status", "Curse", "Quest", "Unknown"]

POOL_NAMES = {
    "Silent": "静默猎手",
    "Colorless": "无色",
    "Ironclad": "铁甲战士",
    "Defect": "故障机器人",
    "Necrobinder": "死灵法师",
    "Regent": "摄政王",
    "Curse": "诅咒",
    "Status": "状态",
    "Event": "事件",
    "Token": "代币",
    "Quest": "任务",
    "Unknown": "未知"
}

POOL_ORDER = ["Silent", "Colorless", "Ironclad", "Defect", "Necrobinder", "Regent", "Curse", "Status", "Event", "Token", "Quest", "Unknown"]

COLOR_STYLES = {
    "blue": {"name": "蓝色", "color": "CCE5FF"},
    "green": {"name": "绿色", "color": "C6EFCE"},
    "orange": {"name": "橙色", "color": "FCE4D6"},
    "yellow": {"name": "黄色", "color": "FFEB9C"},
    "purple": {"name": "紫色", "color": "E4DFEC"},
    "gray": {"name": "灰色", "color": "D9D9D9"},
    "pink": {"name": "粉色", "color": "FFC7CE"},
    "red": {"name": "红色", "color": "F4CCCC"},
    "cyan": {"name": "青色", "color": "D9EAD3"},
}

def load_cards_localization(localization_file):
    if localization_file and Path(localization_file).exists():
        with open(localization_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        card_names = {}
        for key, value in data.items():
            if key.endswith('.title'):
                card_id = key.replace('.title', '')
                card_names[card_id] = value
        return card_names
    return {}

def load_card_info(card_info_file):
    if Path(card_info_file).exists():
        with open(card_info_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def get_run_files(history_dir):
    run_files = list(Path(history_dir).glob("*.run"))
    run_files.sort(key=lambda x: int(x.stem), reverse=True)
    return run_files

def is_character_run(run_data, character="SILENT"):
    for player in run_data.get("players", []):
        if player.get("character") == f"CHARACTER.{character}":
            return True
    return False

def analyze_deck_composition(run_data, card_info):
    deck_stats = {
        "total": 0,
        "curse": 0,
        "strike": 0,
        "defend": 0,
        "attack": 0,
        "skill": 0,
        "power": 0,
        "exhaust": 0,
        "note": ""
    }
    
    has_ghost_seed = False
    has_paels_wing = False
    has_purity = False
    purity_upgraded = False
    has_protector = False
    souls_power_count = 0
    
    for player in run_data.get("players", []):
        deck = player.get("deck", [])
        for card in deck:
            card_id = card.get("id", "").replace("CARD.", "")
            floor_added = card.get("floor_added_to_deck", 0)
            
            deck_stats["total"] += 1
            
            info = card_info.get(card_id, {})
            card_type = info.get("type", "Unknown") if isinstance(info, dict) else "Unknown"
            has_exhaust = info.get("exhaust", False) if isinstance(info, dict) else False
            has_ethereal = info.get("ethereal", False) if isinstance(info, dict) else False
            
            enchantment = card.get("enchantment", {})
            enchantment_id = enchantment.get("id", "").replace("ENCHANTMENT.", "") if enchantment else ""
            
            if has_exhaust or has_ethereal:
                if enchantment_id == "SOULS_POWER":
                    souls_power_count += 1
                else:
                    deck_stats["exhaust"] += 1
            
            if card_id == "PURITY":
                has_purity = True
                purity_upgraded = card.get("current_upgrade_level", 0) > 0
            elif card_id == "PROTECTOR":
                has_protector = True
            
            if card_type == "Curse":
                deck_stats["curse"] += 1
            elif floor_added == 1:
                if "STRIKE" in card_id:
                    deck_stats["strike"] += 1
                elif "DEFEND" in card_id:
                    deck_stats["defend"] += 1
                else:
                    deck_stats["skill"] += 1
            elif card_type == "Attack":
                deck_stats["attack"] += 1
            elif card_type == "Skill":
                deck_stats["skill"] += 1
            elif card_type == "Power":
                deck_stats["power"] += 1
        
        relics = player.get("relics", [])
        for relic in relics:
            relic_id = relic.get("id", "").replace("RELIC.", "")
            if relic_id == "GHOST_SEED":
                has_ghost_seed = True
            elif relic_id == "PAELS_WING":
                has_paels_wing = True
    
    net = deck_stats["total"] - deck_stats["exhaust"] - deck_stats["power"]
    notes = []
    
    if souls_power_count > 0:
        net += souls_power_count
    
    if has_ghost_seed:
        net -= deck_stats["strike"] + deck_stats["defend"]
        notes.append(f"幽灵种子(-{deck_stats['strike']+deck_stats['defend']})")
    
    if has_paels_wing:
        net -= deck_stats["defend"]
        notes.append(f"佩尔之翼(-{deck_stats['defend']})")
    
    if has_purity:
        purity_amount = 5 if purity_upgraded else 3
        if net > purity_amount:
            net -= purity_amount
            notes.append(f"净化(-{purity_amount})")
        else:
            notes.append(f"净化(净数量≤{purity_amount},不调整)")
    
    if has_protector:
        if net <= 20:
            net = 11
            notes.append("护驾！！！(净数量≤20,调整为11)")
        else:
            notes.append("护驾！！！(净数量>20,不调整)")
    
    deck_stats["net"] = net
    deck_stats["note"] = "; ".join(notes) if notes else ""
    
    return deck_stats

def style_header(ws, num_cols, has_merge=False, color="CCE5FF"):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_num in range(1, 3 if has_merge else 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

def create_rarity_sheet(ws, card_stats, total_stats, cards_by_rarity, card_names, rarity, color="CCE5FF"):
    cards = cards_by_rarity.get(rarity, [])
    if not cards:
        return 0
    
    header_row1 = ["卡牌ID", "卡牌", "第一幕", "", "", "第二幕", "", "", "第三幕", "", "", "总计", "", ""]
    header_row2 = ["", "", "抓取数", "出现数", "抓取率", "抓取数", "出现数", "抓取率", "抓取数", "出现数", "抓取率", "抓取数", "出现数", "抓取率"]
    
    ws.append(header_row1)
    ws.append(header_row2)
    
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:E1')
    ws.merge_cells('F1:H1')
    ws.merge_cells('I1:K1')
    ws.merge_cells('L1:N1')
    
    sorted_cards = sorted(
        cards,
        key=lambda x: total_stats[x]["picked"] / total_stats[x]["offered"] if total_stats[x]["offered"] > 0 else 0,
        reverse=True
    )
    
    for card_id in sorted_cards:
        cn_name = card_names.get(card_id, "")
        row = [card_id, cn_name]
        
        for act in [0, 1, 2]:
            stats = card_stats[act].get(card_id, {"picked": 0, "offered": 0})
            picked = stats["picked"]
            offered = stats["offered"]
            pick_rate = picked / offered if offered > 0 else 0
            row.extend([picked, offered, pick_rate])
        
        total = total_stats[card_id]
        total_picked = total["picked"]
        total_offered = total["offered"]
        total_rate = total_picked / total_offered if total_offered > 0 else 0
        row.extend([total_picked, total_offered, total_rate])
        
        ws.append(row)
    
    style_header(ws, len(header_row1), has_merge=True, color=color)
    
    for row_idx in range(3, ws.max_row + 1):
        for col_idx in [3, 4, 6, 7, 9, 10, 12, 13]:
            ws.cell(row=row_idx, column=col_idx).number_format = '0'
        for col_idx in [5, 8, 11, 14]:
            ws.cell(row=row_idx, column=col_idx).number_format = '0.0%'
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 10
    
    return len(cards)

def analyze_runs(history_dir, card_info_file, n, m, output_dir, localization_file=None, style="blue", character="SILENT"):
    color = COLOR_STYLES.get(style, COLOR_STYLES["blue"])["color"]
    style_name = COLOR_STYLES.get(style, COLOR_STYLES["blue"])["name"]
    
    print(f"正在分析倒数第{n}到第{m}局{character}游戏...")
    print(f"颜色风格: {style_name}")
    print("=" * 60)
    
    run_files = get_run_files(history_dir)
    print(f"共找到 {len(run_files)} 个历史记录文件")
    
    character_runs = []
    for run_file in run_files:
        try:
            with open(run_file, 'r', encoding='utf-8') as f:
                run_data = json.load(f)
            if is_character_run(run_data, character):
                character_runs.append((run_file, run_data))
        except Exception as e:
            print(f"读取文件 {run_file} 时出错: {e}")
    
    print(f"其中{character}游戏 {len(character_runs)} 局")
    
    total_runs = len(character_runs)
    if m > total_runs:
        print(f"警告: m值({m})超过游戏总数({total_runs})")
        m = total_runs
    if n > total_runs:
        print(f"警告: n值({n})超过游戏总数({total_runs})")
        n = total_runs
    
    start_idx = n - 1
    end_idx = m
    
    selected_runs = character_runs[start_idx:end_idx]
    print(f"选择倒数第{n}到第{m}局，共 {len(selected_runs)} 局")
    print("=" * 60)
    
    card_stats = {
        0: defaultdict(lambda: {"picked": 0, "offered": 0}),
        1: defaultdict(lambda: {"picked": 0, "offered": 0}),
        2: defaultdict(lambda: {"picked": 0, "offered": 0}),
    }
    
    deck_compositions = []
    version_stats = defaultdict(int)
    
    card_info = load_card_info(card_info_file)
    
    for i, (run_file, run_data) in enumerate(selected_runs):
        actual_idx = start_idx + i + 1
        build_id = run_data.get("build_id", "Unknown")
        version_stats[build_id] += 1
        print(f"\n处理倒数第{actual_idx}局: {run_file.name} (版本: {build_id})")
        
        deck_comp = analyze_deck_composition(run_data, card_info)
        deck_compositions.append(deck_comp)
        print(f"  卡组构成: 总{deck_comp['total']}张, 诅咒{deck_comp['curse']}, 打击{deck_comp['strike']}, 防御{deck_comp['defend']}, 攻击{deck_comp['attack']}, 技能{deck_comp['skill']}, 能力{deck_comp['power']}, 消耗{deck_comp['exhaust']}, 净{deck_comp['net']}张" + (f" [{deck_comp['note']}]" if deck_comp['note'] else ""))
        
        map_history = run_data.get("map_point_history", [])
        for act_index, act_points in enumerate(map_history):
            if act_index > 2:
                break
            
            for point in act_points:
                rooms = point.get("rooms", [])
                room_type = rooms[0].get("room_type", "unknown") if rooms else "unknown"
                
                player_stats = point.get("player_stats", [])
                for stats in player_stats:
                    card_choices = stats.get("card_choices", [])
                    for choice in card_choices:
                        card = choice.get("card", {})
                        card_id = card.get("id", "")
                        if card_id:
                            card_id_clean = card_id.replace("CARD.", "")
                            card_stats[act_index][card_id_clean]["offered"] += 1
                            if choice.get("was_picked", False):
                                card_stats[act_index][card_id_clean]["picked"] += 1
                    
                    if room_type == "shop":
                        cards_gained = stats.get("cards_gained", [])
                        for gained in cards_gained:
                            card_id = gained.get("id", "")
                            if card_id:
                                card_id_clean = card_id.replace("CARD.", "")
                                card_stats[act_index][card_id_clean]["offered"] += 1
                                card_stats[act_index][card_id_clean]["picked"] += 1
    
    card_names = load_cards_localization(localization_file)
    
    print("\n" + "=" * 60)
    print("卡组构成统计")
    print("=" * 60)
    
    num_runs = len(deck_compositions)
    avg_stats = {
        "total": sum(d["total"] for d in deck_compositions) / num_runs,
        "curse": sum(d["curse"] for d in deck_compositions) / num_runs,
        "strike": sum(d["strike"] for d in deck_compositions) / num_runs,
        "defend": sum(d["defend"] for d in deck_compositions) / num_runs,
        "attack": sum(d["attack"] for d in deck_compositions) / num_runs,
        "skill": sum(d["skill"] for d in deck_compositions) / num_runs,
        "power": sum(d["power"] for d in deck_compositions) / num_runs,
        "exhaust": sum(d["exhaust"] for d in deck_compositions) / num_runs,
        "net": sum(d["net"] for d in deck_compositions) / num_runs,
    }
    
    print(f"统计局数: {num_runs}")
    print(f"版本分布:")
    for version, count in sorted(version_stats.items(), reverse=True):
        print(f"  {version}: {count}局")
    print(f"平均卡组构成:")
    print(f"  总数量: {avg_stats['total']:.1f}")
    print(f"  诅咒: {avg_stats['curse']:.1f}")
    print(f"  打击: {avg_stats['strike']:.1f}")
    print(f"  防御: {avg_stats['defend']:.1f}")
    print(f"  攻击牌(非初始): {avg_stats['attack']:.1f}")
    print(f"  技能牌(非初始): {avg_stats['skill']:.1f}")
    print(f"  能力牌: {avg_stats['power']:.1f}")
    print(f"  消耗牌: {avg_stats['exhaust']:.1f}")
    print(f"  净数量: {avg_stats['net']:.1f}")
    
    print("\n" + "=" * 60)
    print("卡牌抓取统计")
    print("=" * 60)
    
    all_cards = set()
    for act in [0, 1, 2]:
        all_cards.update(card_stats[act].keys())
    
    total_stats = defaultdict(lambda: {"picked": 0, "offered": 0})
    for act in [0, 1, 2]:
        for card_id, stats in card_stats[act].items():
            total_stats[card_id]["picked"] += stats["picked"]
            total_stats[card_id]["offered"] += stats["offered"]
    
    cards_by_pool_rarity = defaultdict(lambda: defaultdict(list))
    for card_id in all_cards:
        info = card_info.get(card_id, {})
        pool = info.get("pool", "Unknown") if isinstance(info, dict) else "Unknown"
        rarity = info.get("rarity", "Unknown") if isinstance(info, dict) else "Unknown"
        cards_by_pool_rarity[pool][rarity].append(card_id)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    output_file = Path(output_dir) / f"card_stats_{n}_{m}_{timestamp}.xlsx"
    
    wb = Workbook()
    ws_deck = wb.active
    ws_deck.title = "卡组构成"
    
    deck_header = ["局数", "总数量", "诅咒", "打击\n(初始)", "防御\n(初始)", "攻击牌\n(非初始)", "技能牌\n(非初始)", "能力牌", "消耗牌", "净数量", "备注"]
    ws_deck.append(deck_header)
    
    for i, deck_comp in enumerate(deck_compositions):
        row = [start_idx + i + 1, deck_comp["total"], deck_comp["curse"], 
               deck_comp["strike"], deck_comp["defend"], deck_comp["attack"], deck_comp["skill"], deck_comp["power"], deck_comp["exhaust"], deck_comp["net"], deck_comp["note"]]
        ws_deck.append(row)
    
    avg_row = len(deck_compositions) + 2
    ws_deck.append(["平均", avg_stats['total'], avg_stats['curse'],
                    avg_stats['strike'], avg_stats['defend'],
                    avg_stats['attack'], avg_stats['skill'], avg_stats['power'], avg_stats['exhaust'], avg_stats['net'], ""])
    
    ws_deck.append([])
    ws_deck.append(["版本分布"])
    for version, count in sorted(version_stats.items(), reverse=True):
        ws_deck.append([version, f"{count}局"])
    
    for row_idx in range(2, avg_row):
        for col_idx in range(2, 11):
            cell = ws_deck.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0'
    
    for col_idx in range(2, 11):
        cell = ws_deck.cell(row=avg_row, column=col_idx)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.0'
            cell.font = Font(bold=True)
    
    style_header(ws_deck, len(deck_header), color=color)
    
    ws_deck.column_dimensions['A'].width = 8
    ws_deck.column_dimensions['B'].width = 10
    ws_deck.column_dimensions['C'].width = 8
    ws_deck.column_dimensions['D'].width = 10
    ws_deck.column_dimensions['E'].width = 10
    ws_deck.column_dimensions['F'].width = 10
    ws_deck.column_dimensions['G'].width = 10
    ws_deck.column_dimensions['H'].width = 10
    ws_deck.column_dimensions['I'].width = 10
    ws_deck.column_dimensions['J'].width = 10
    ws_deck.column_dimensions['K'].width = 40
    
    ws_deck.row_dimensions[1].height = 30
    
    print(f"\n[卡组构成] 已添加到表单")
    
    for pool in POOL_ORDER:
        pool_name = POOL_NAMES.get(pool, pool)
        for rarity in RARITY_ORDER:
            rarity_name = RARITY_NAMES.get(rarity, rarity)
            sheet_name = f"{pool_name}-{rarity_name}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            card_count = create_rarity_sheet(ws, card_stats, total_stats, cards_by_pool_rarity[pool], card_names, rarity, color=color)
            
            if card_count > 0:
                print(f"[{pool_name}-{rarity_name}] {card_count}张卡牌 -> 表单: {sheet_name}")
    
    wb.save(output_file)
    print(f"\n统计结果已保存到: {output_file}")
    return output_file

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="统计Slay The Spire 2历史记录中卡牌的抓取数和抓取率")
    parser.add_argument("n", type=int, help="起始局数（倒数第n局）")
    parser.add_argument("m", type=int, help="结束局数（倒数第m局，m >= n）")
    parser.add_argument("--history", "-i", type=str, required=True, 
                        help="历史记录目录路径（包含.run文件的目录）")
    parser.add_argument("--card-info", "-c", type=str, required=True,
                        help="卡牌信息JSON文件路径（card_rarity.json）")
    parser.add_argument("--output", "-o", type=str, default="card_stats_output",
                        help="输出目录路径（默认: card_stats_output）")
    parser.add_argument("--localization", "-l", type=str, default=None,
                        help="本地化文件路径（可选，用于显示中文卡牌名）")
    parser.add_argument("--style", "-s", type=str, default="blue", 
                        choices=list(COLOR_STYLES.keys()),
                        help="表头颜色风格")
    parser.add_argument("--character", type=str, default="SILENT",
                        help="角色名称（默认: SILENT）")
    args = parser.parse_args()
    
    if args.n < 1:
        print("错误: n必须大于等于1")
        exit(1)
    if args.m < args.n:
        print("错误: m不能小于n")
        exit(1)
    
    analyze_runs(args.history, args.card_info, args.n, args.m, args.output, args.localization, args.style, args.character)
